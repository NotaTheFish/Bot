import unittest
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from accountant_bot.excel_export import build_transactions_report


class AccountantBotExcelExportTests(unittest.TestCase):
    def test_build_transactions_report_includes_styled_localized_sheets_and_summary_net(self):
        transactions = [
            {
                "receipt_id": 2,
                "created_at": "2026-01-01 11:00:00",
                "admin": "User Name (id: 101)",
                "currency": "UAH",
                "pay_method": "card",
                "total_sum": "10",
                "note": "",
                "receipt_file_id": None,
                "status": "refunded",
                "items": [
                    {
                        "category": "TOKENS",
                        "item_name": "Tip",
                        "qty": "2",
                        "unit_price": "20",
                        "unit_basis": "unit",
                        "line_total": "40",
                        "note": "",
                    },
                ],
            },
            {
                "receipt_id": 1,
                "created_at": "2026-01-01 10:00:00",
                "admin": "@admin (id: 100)",
                "currency": "RUB",
                "pay_method": "cash",
                "total_sum": "200",
                "note": "morning",
                "receipt_file_id": "abc",
                "status": "created",
                "items": [
                    {
                        "category": "VID",
                        "item_name": "Coffee",
                        "qty": "2",
                        "unit_price": "100",
                        "unit_basis": "unit",
                        "line_total": "200",
                        "note": "",
                    },
                    {
                        "category": "OTHER",
                        "item_name": "Pack",
                        "qty": "1",
                        "unit_price": "7",
                        "unit_basis": "unit",
                        "line_total": "7",
                        "note": "",
                    },

                    {
                        "category": "COINS",
                        "item_name": "Coins pack",
                        "qty": "1",
                        "unit_price": "150",
                        "unit_basis": "per_100000",
                        "line_total": "150",
                        "note": "",
                    },
                ],
            },
        ]

        report = build_transactions_report(transactions, ZoneInfo("Asia/Tokyo"))

        wb = load_workbook(BytesIO(report))
        self.assertEqual(wb.sheetnames, ["Чеки", "Позиции", "Сводка"])

        ws_receipts = wb["Чеки"]
        self.assertEqual(ws_receipts["A2"].value, 2)  # sorted by created_at DESC
        self.assertEqual(ws_receipts["I2"].value, "REFUND")
        self.assertEqual(ws_receipts["B2"].value, datetime(2026, 1, 1, 20, 0))
        self.assertEqual(ws_receipts["B3"].value, datetime(2026, 1, 1, 19, 0))
        self.assertEqual(ws_receipts.freeze_panes, "A2")

        ws_items = wb["Позиции"]
        self.assertEqual(ws_items["F2"].value, "🐉 Вид")
        self.assertEqual(ws_items["F3"].value, "🪙 Токены")
        self.assertEqual(ws_items["F4"].value, "✍️ Другое")
        self.assertEqual(ws_items["F5"].value, "🪙 Coins")
        self.assertEqual(ws_items["J2"].value, "шт")
        self.assertEqual(ws_items["J3"].value, "шт")
        self.assertEqual(ws_items["J5"].value, "за 100000")

        ws_summary = wb["Сводка"]
        self.assertEqual(ws_summary["B1"].value, "Timezone: Asia/Tokyo")
        self.assertTrue(str(ws_summary["C1"].value).startswith("Generated at:"))
        self.assertEqual(ws_summary["A4"].value, "Валюта")
        self.assertEqual(ws_summary["A5"].value, "RUB")
        self.assertEqual(ws_summary["D5"].value, 357)
        self.assertEqual(ws_summary["A6"].value, "UAH")
        self.assertEqual(ws_summary["D6"].value, -40)
        self.assertEqual(ws_summary["A10"].value, "Доход по дням")
        self.assertEqual(ws_summary["A16"].value, "Топ товаров")
        self.assertIn("Доход по категориям", [cell.value for cell in ws_summary["A"]])
        self.assertEqual(ws_summary.freeze_panes, "A4")
        self.assertEqual(len(ws_summary._charts), 2)

        chart_1_anchor = ws_summary._charts[0].anchor._from
        chart_2_anchor = ws_summary._charts[1].anchor._from
        self.assertEqual((chart_1_anchor.col, chart_1_anchor.row), (5, 2))
        self.assertEqual((chart_2_anchor.col, chart_2_anchor.row), (5, 19))


if __name__ == "__main__":
    unittest.main()