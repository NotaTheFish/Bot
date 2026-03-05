from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .category_labels import CATEGORY_LABELS

RECEIPTS_HEADERS = [
    "receipt_id",
    "created_at",
    "admin",
    "currency",
    "pay_method",
    "total_sum",
    "note",
    "has_receipt",
    "status",
]

ITEMS_HEADERS = [
    "receipt_id",
    "created_at",
    "admin",
    "status",
    "currency",
    "category",
    "item_name",
    "qty",
    "unit_price",
    "unit_basis",
    "line_total",
    "note",
]

UNIT_BASIS_LABELS = {
    "unit": "шт",
    "per_1000": "за 1000",
}

CATEGORY_ORDER = {"VID": 1, "TOKENS": 2, "MUSHROOMS": 3, "OTHER": 4}
CATEGORY_FILLS = {
    "VID": PatternFill("solid", fgColor="FFDCEBFF"),
    "TOKENS": PatternFill("solid", fgColor="FFE8F5E9"),
    "MUSHROOMS": PatternFill("solid", fgColor="FFFCE4EC"),
    "OTHER": PatternFill("solid", fgColor="FFFFF3E0"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="FFE5E7EB"),
    right=Side(style="thin", color="FFE5E7EB"),
    top=Side(style="thin", color="FFE5E7EB"),
    bottom=Side(style="thin", color="FFE5E7EB"),
)
HEADER_FILL = PatternFill("solid", fgColor="FFF3F4F6")


def build_transactions_report(transactions: list[dict[str, Any] | Any], tz: ZoneInfo) -> bytes:
    receipts: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []

    for row in transactions:
        record = dict(row)
        created_at = _coerce_datetime(record.get("created_at"))
        status = _normalize_status(record.get("status"))
        receipt = {
            "receipt_id": record.get("receipt_id") or record.get("id"),
            "created_at": created_at,
            "admin": str(record.get("admin") or record.get("admin_id") or ""),
            "currency": str(record.get("currency") or ""),
            "pay_method": str(record.get("pay_method") or ""),
            "total_sum": _to_decimal(record.get("total_sum") if record.get("total_sum") is not None else record.get("total")),
            "note": str(record.get("note") or ""),
            "has_receipt": bool(record.get("receipt_file_id")),
            "status": status,
        }
        receipts.append(receipt)

        for item in record.get("items") or []:
            item_record = dict(item)
            category_code = str(item_record.get("category") or "").upper() or "OTHER"
            items.append(
                {
                    "receipt_id": receipt["receipt_id"],
                    "created_at": created_at,
                    "admin": receipt["admin"],
                    "status": status,
                    "currency": receipt["currency"],
                    "category_code": category_code,
                    "category": CATEGORY_LABELS.get(category_code, category_code),
                    "item_name": str(item_record.get("item_name") or ""),
                    "qty": _to_decimal(item_record.get("qty")),
                    "unit_price": _to_decimal(item_record.get("unit_price")),
                    "unit_basis": UNIT_BASIS_LABELS.get(str(item_record.get("unit_basis") or ""), str(item_record.get("unit_basis") or "")),
                    "line_total": _to_decimal(item_record.get("line_total")),
                    "note": str(item_record.get("note") or ""),
                }
            )

    receipts.sort(key=lambda r: (_sortable_dt(r["created_at"]), int(r.get("receipt_id") or 0)), reverse=True)
    items.sort(
        key=lambda i: (
            CATEGORY_ORDER.get(i.get("category_code") or "OTHER", 99),
            str(i.get("item_name") or "").lower(),
            int(i.get("receipt_id") or 0),
        )
    )

    wb = Workbook()
    ws_receipts = wb.active
    ws_receipts.title = "Чеки"
    _build_receipts_sheet(ws_receipts, receipts, tz)

    ws_items = wb.create_sheet("Позиции")
    _build_items_sheet(ws_items, items, tz)

    ws_summary = wb.create_sheet("Сводка")
    _build_summary_sheet(ws_summary, items, tz)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_receipts_sheet(ws: Any, receipts: list[dict[str, Any]], tz: ZoneInfo) -> None:
    ws.append(RECEIPTS_HEADERS)
    for receipt in receipts:
        ws.append(
            [
                receipt["receipt_id"],
                _to_excel_dt(receipt["created_at"], tz),
                receipt["admin"],
                receipt["currency"],
                receipt["pay_method"],
                receipt["total_sum"],
                receipt["note"],
                "yes" if receipt["has_receipt"] else "no",
                receipt["status"],
            ]
        )
    _apply_sheet_style(ws, money_columns={6}, datetime_columns={2})
    ws.auto_filter.ref = ws.dimensions


def _build_items_sheet(ws: Any, items: list[dict[str, Any]], tz: ZoneInfo) -> None:
    ws.append(ITEMS_HEADERS)
    for item in items:
        ws.append(
            [
                item["receipt_id"],
                _to_excel_dt(item["created_at"], tz),
                item["admin"],
                item["status"],
                item["currency"],
                item["category"],
                item["item_name"],
                item["qty"],
                item["unit_price"],
                item["unit_basis"],
                item["line_total"],
                item["note"],
            ]
        )

    _apply_sheet_style(ws, money_columns={9, 11}, qty_columns={8}, datetime_columns={2})
    ws.auto_filter.ref = ws.dimensions

    for row_idx, item in enumerate(items, start=2):
        fill = category_fill(item.get("category_code") or item.get("category") or "")
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _build_summary_sheet(ws: Any, items: list[dict[str, Any]], tz: ZoneInfo) -> None:
    ws["A1"] = "Сводка по транзакциям"
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"] = f"Timezone: {tz.key}"
    ws["C1"] = f"Generated at: {datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')}"

    currencies = sorted({(item.get("currency") or "-") for item in items} or {"-"})

    row_cursor = 3

    totals: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"ok": Decimal("0"), "refund": Decimal("0")})
    category_totals: dict[str, dict[str, dict[str, Decimal]]] = defaultdict(lambda: defaultdict(lambda: {"ok": Decimal("0"), "refund": Decimal("0")}))
    day_totals: dict[datetime.date, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    product_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"qty": Decimal("0"), "currency_net": defaultdict(Decimal)})

    for item in items:
        status = item.get("status")
        if status == "CANCELED":
            continue

        currency = item.get("currency") or "-"
        amount = _to_decimal(item.get("line_total"))
        factor = _status_factor(status)

        if status == "REFUND":
            totals[currency]["refund"] += amount
        else:
            totals[currency]["ok"] += amount

        category_code = item.get("category_code") or "OTHER"
        if status == "REFUND":
            category_totals[category_code][currency]["refund"] += amount
        else:
            category_totals[category_code][currency]["ok"] += amount

        created_at = item.get("created_at")
        if isinstance(created_at, datetime):
            local_created_at = _to_excel_dt(created_at, tz)
            if isinstance(local_created_at, datetime):
                day_totals[local_created_at.date()][currency] += factor * amount

        product_name = str(item.get("item_name") or "-")
        product_totals[product_name]["qty"] += _to_decimal(item.get("qty"))
        product_totals[product_name]["currency_net"][currency] += factor * amount

    ws.cell(row=row_cursor, column=1, value="Общий доход по валютам")
    ws.cell(row=row_cursor, column=1).font = Font(bold=True)
    summary_header_row = row_cursor + 1
    ws.append(["Валюта", "OK", "REFUND", "NET"])

    for currency in currencies:
        ok_sum = totals[currency]["ok"]
        refund_sum = totals[currency]["refund"]
        ws.append([currency, ok_sum, refund_sum, ok_sum - refund_sum])
    summary_end = ws.max_row
    _style_table(ws, summary_header_row, summary_end, money_columns={2, 3, 4})
    row_cursor = summary_end + 4

    ws.cell(row=row_cursor, column=1, value="Доход по дням")
    ws.cell(row=row_cursor, column=1).font = Font(bold=True)
    day_header_row = row_cursor + 1
    day_headers = ["Дата", *[f"{currency}_NET" for currency in currencies]]
    ws.append(day_headers)
    for day in sorted(day_totals):
        ws.append([day, *[day_totals[day][currency] for currency in currencies]])
    day_end = ws.max_row
    _style_table(ws, day_header_row, day_end, money_columns=set(range(2, len(day_headers) + 1)), date_columns={1})
    row_cursor = day_end + 4

    ws.cell(row=row_cursor, column=1, value="Топ товаров")
    ws.cell(row=row_cursor, column=1).font = Font(bold=True)
    top_header_row = row_cursor + 1
    top_headers = ["Товар", "Продано (qty)", *[f"{currency}_NET" for currency in currencies]]
    ws.append(top_headers)

    sorted_products = sorted(
        product_totals.items(),
        key=lambda pair: sum(pair[1]["currency_net"].values()),
        reverse=True,
    )
    for product_name, agg in sorted_products[:10]:
        ws.append([product_name, agg["qty"], *[agg["currency_net"][currency] for currency in currencies]])
    top_end = ws.max_row
    _style_table(ws, top_header_row, top_end, money_columns=set(range(3, len(top_headers) + 1)), qty_columns={2})
    row_cursor = top_end + 4

    ws.cell(row=row_cursor, column=1, value="Доход по категориям")
    ws.cell(row=row_cursor, column=1).font = Font(bold=True)
    category_header_row = row_cursor + 1

    category_headers = ["Категория"]
    net_columns: list[int] = []
    for currency in currencies:
        category_headers.extend([f"{currency}_OK", f"{currency}_REFUND", f"{currency}_NET"])
        net_columns.append(len(category_headers))
    ws.append(category_headers)

    for category_code, _ in sorted(CATEGORY_ORDER.items(), key=lambda pair: pair[1]):
        category_label = CATEGORY_LABELS.get(category_code, category_code)
        row_values: list[Any] = [category_label]
        for currency in currencies:
            data = category_totals[category_code][currency]
            row_values.extend([data["ok"], data["refund"], data["ok"] - data["refund"]])
        ws.append(row_values)
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = category_fill(category_code)

    category_end = ws.max_row
    money_cols = set(range(2, len(category_headers) + 1))
    _style_table(ws, category_header_row, category_end, money_columns=money_cols)

    _add_summary_charts(
        ws,
        day_header_row,
        day_end,
        category_header_row,
        category_end,
        len(currencies),
        net_columns,
    )

    ws.freeze_panes = "A4"
    autosize_columns(ws)


def _add_summary_charts(
    ws: Any,
    day_header_row: int,
    day_end: int,
    category_header_row: int,
    category_end: int,
    currencies_count: int,
    category_net_columns: list[int],
) -> None:
    chart_column = "F"
    chart_1_row = 3
    chart_height_rows = 14
    chart_gap_rows = 3

    if day_end > day_header_row:
        line_chart = LineChart()
        line_chart.title = "Доход по дням"
        line_chart.style = 10
        line_chart.y_axis.title = "NET"
        line_chart.x_axis.title = "Дата"

        day_data = Reference(
            ws,
            min_col=2,
            max_col=1 + currencies_count,
            min_row=day_header_row,
            max_row=day_end,
        )
        day_labels = Reference(ws, min_col=1, min_row=day_header_row + 1, max_row=day_end)
        line_chart.add_data(day_data, titles_from_data=True)
        line_chart.set_categories(day_labels)
        line_chart.height = 7
        line_chart.width = 14
        ws.add_chart(line_chart, f"{chart_column}{chart_1_row}")

    if category_end > category_header_row:
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Доход по категориям"
        bar_chart.y_axis.title = "NET"

        for col_idx in category_net_columns:
            data = Reference(ws, min_col=col_idx, max_col=col_idx, min_row=category_header_row, max_row=category_end)
            bar_chart.add_data(data, titles_from_data=True)
        labels = Reference(ws, min_col=1, min_row=category_header_row + 1, max_row=category_end)
        bar_chart.set_categories(labels)
        bar_chart.height = 7
        bar_chart.width = 14
        chart_2_row = chart_1_row + chart_height_rows + chart_gap_rows
        ws.add_chart(bar_chart, f"{chart_column}{chart_2_row}")


def _apply_sheet_style(
    ws: Any,
    *,
    money_columns: set[int],
    qty_columns: set[int] | None = None,
    datetime_columns: set[int] | None = None,
) -> None:
    qty_columns = qty_columns or set()
    datetime_columns = datetime_columns or set()
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in money_columns:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for each in cell:
                each.number_format = "#,##0.00"

    for col_idx in qty_columns:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for each in cell:
                each.number_format = "#,##0.###"

    for col_idx in datetime_columns:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for each in cell:
                each.number_format = "DD.MM.YYYY HH:MM"

    autosize_columns(ws)


def _style_table(
    ws: Any,
    header_row: int,
    end_row: int,
    *,
    money_columns: set[int] | None = None,
    qty_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
) -> None:
    money_columns = money_columns or set()
    qty_columns = qty_columns or set()
    date_columns = date_columns or set()

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx in range(header_row + 1, end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col_idx in money_columns:
                cell.number_format = "#,##0.00"
            elif col_idx in qty_columns:
                cell.number_format = "#,##0.###"
            elif col_idx in date_columns:
                cell.number_format = "DD.MM.YYYY"


def autosize_columns(ws: Any) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column_letter].width = min(max(12, max_len + 2), 60)


def category_fill(category_label_or_code: str) -> PatternFill:
    normalized = str(category_label_or_code or "").strip()
    if normalized in CATEGORY_FILLS:
        return CATEGORY_FILLS[normalized]
    for code, label in CATEGORY_LABELS.items():
        if normalized == label:
            return CATEGORY_FILLS[code]
    return CATEGORY_FILLS["OTHER"]


def _normalize_status(value: Any) -> str:
    status_raw = str(value or "").strip().lower()
    if status_raw in {"refunded", "refund"}:
        return "REFUND"
    if status_raw in {"canceled", "cancelled", "cancel"}:
        return "CANCELED"
    return "OK"


def _status_factor(status: Any) -> Decimal:
    normalized = _normalize_status(status)
    if normalized == "REFUND":
        return Decimal("-1")
    if normalized == "CANCELED":
        return Decimal("0")
    return Decimal("1")


def _sortable_dt(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.min


def _coerce_datetime(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return None
    raw = str(value).strip()
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return raw


def _to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_excel_dt(value: Any, tz: ZoneInfo) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(tz).replace(tzinfo=None)
        return value.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
    return value